"""Normalize MFD per-sample habitat metadata into one barcode-keyed table.

Prototype normalizer. Ingests every file in the 39 project folders + the
habitat ontology from the cmc-aau/mfd_metadata repo (pinned commit), reconciles
the heterogeneous per-project schemas into a single table keyed by
`fieldsample_barcode`, and reports joinable coverage against the PRJNA1071982
NMDC biosamples.

The motivation: `isolation_source` on the NCBI biosamples is a lossy MFD
Level-1 projection ("Soil from Natural Dunes"), which is *why* the env-triad
curation leaves ~2,580 sentinels. Every NMDC biosample already carries an
`MFDID` (== `fieldsample_barcode`), which joins to the full per-sample
hab1/hab2/hab3 classification here. This script proves out that join.

Habitat-signal tiers, best-first:
  - standard  : file already has mfd_hab1/hab2/hab3 columns (with values)
  - natura2000: file has a Natura-2000 / MFD habitat code (Naturnr,
                habitat_typenumber, hab_code) -> resolved via the MFD ontology
  - unified   : agriculture file has a `unified` code (so_a_NNNN) whose numeric
                suffix is an MFD habitat code -> resolved via the ontology
  - sampletype: only soil/water/sediment + natural/urban/... is recoverable

Inputs are pinned to mfd_metadata commit b1b17d4 (2026-02-23).

Requires: openpyxl (stdlib otherwise).

Writes:
  - data/mfd_habitat_per_sample.tsv          (the normalized table)
  - results/mfd_habitat_coverage_report.json (the coverage report)
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

REF = "b1b17d4b4e8e289c4380d878c8bf1516eb107791"  # mfd_metadata pinned commit
RAW = f"https://raw.githubusercontent.com/cmc-aau/mfd_metadata/{REF}/data"

REPO = Path(__file__).resolve().parent.parent
CACHE = REPO / "data" / "mfd_metadata_cache"
OUT_TSV = REPO / "data" / "mfd_habitat_per_sample.tsv"
OUT_REPORT = REPO / "results" / "mfd_habitat_coverage_report.json"
CURATION_INPUTS = REPO / "results" / "ncbi_PRJNA1071982_nmdc_curation_inputs.json"
CURATION_REPORT = REPO / "results" / "ncbi_PRJNA1071982_nmdc_curation_report.json"

ONTOLOGY_PATH = "ontology/latest_mfd-habitat-ontology.xlsx"
NCBI_META_PATH = "metadata/general/2025-04-14_MFD_NCBI_meta 1.tsv"

# Every habitat-relevant file in each project folder. The parser is fully
# schema-driven (it inspects headers), so files are listed exhaustively and
# self-filter: a file with no barcode column is silently skipped.
PROJECT_FILES = {
    "P01_1": ["MFDP005_metadata_32.xlsx", "P01_1_filled_table.xlsx",
              "P01_1_filled_table_TGF.xlsx"],
    "P01_2": ["P01_2_MFD metadata_32.xlsx", "P01_2_filled_table.xlsx"],
    "P02_1": ["P02_1_MFD metadata_32.xlsx", "P02_1_filled_table.xlsx"],
    "P02_2": ["P02_2_MFD metadata_32.xlsx", "P02_2_filled_table.xlsx"],
    "P03_1": ["P03_1_MFD metadata_32.xlsx", "P03_1_filled_table.xlsx"],
    "P04_1": ["MFDP006_MFD metadata.xlsx"],
    "P04_2": ["P04_2_filled_table.xlsx"],
    "P04_3": ["2022-11-28_mfd-metadata-kvnet.xlsx", "Kvnet_prøvedatoer 2020.xlsx",
              "Kvnet_prøvedatoer 2021.xlsx", "mapping.xlsx"],
    "P04_4": ["P04_4_filled_table.xlsx"],
    "P04_5": ["2022-11-28_mfd-metadata-seges.xlsx", "centroids_seges.xlsx"],
    "P04_6": ["P04_6_filled_table.xlsx"],
    "P04_7": ["P04_7_filled_table.xlsx"],
    "P04_8": ["P04_8_frde_standard table.xlsx"],
    "P06_2": ["P06_2_frde_standard table.xlsx"],
    "P06_3": ["P06_3_filled_table.xlsx"],
    "P08_1": ["P08_1_filled_table.xlsx"],
    "P08_2": ["Correct coordinates.xlsx", "P08_2_filled_table.xlsx"],
    "P08_3": ["P08_3_filled_table.xlsx"],
    "P08_5": ["2022-07-08_mfd-metadata-sustainscapes.xlsx",
              "230524_SustainScapes_MFD_HabitatClasses_ForVibeke.txt"],
    "P08_6": ["P08_6_filled_table.xlsx", "P08_6_minimal_metadata_batch1&2.xlsx"],
    "P08_7": ["2021-12-22_P08_7-metadata.xlsx"],
    "P08_8": ["P08_8_frde_standard table.xlsx"],
    "P09_1": ["P09_1_filled_table.xlsx"],
    "P09_3": ["P09_3_frde_standard table.xlsx"],
    "P09_4": ["P09_4_frde_standard table.xlsx"],
    "P10_2": ["P10_2_filled_table.xlsx"],
    "P11_2": ["2023-04-25_P11_2-mapping.xlsx", "P11_2_filled_table.xlsx"],
    "P12_1": ["P12_1_minimal_metadata_updated.xlsx"],
    "P12_4": ["P12_4_filled_table.xlsx"],
    "P12_5": ["P12_5_frde_standard table.xlsx"],
    "P13_2": ["P13_2_table_to_fill_Wextrasamples2.xlsx"],
    "P13_3": ["P13_3_frde_standard table1_chj_20230807.xlsx"],
    "P16_2": ["P16_2_filled_table.xlsx"],
    "P16_4": ["P16_4_frde_standard table.xlsx"],
    "P18_2": ["P18_2_frde_standard table.xlsx"],
    "P19_1": ["P19_1_filled_table.xlsx"],
    "P20_1": ["P20_1_frde_standard table.xlsx"],
    "P21_1": ["P21_1_frde_standard table.xlsx"],
    "P25_1": ["P25_1_frde_standard table.xlsx"],
}

BARCODE_RE = re.compile(r"^MFD\d{4,6}$")
SIGNAL_RANK = {"standard": 4, "natura2000": 3, "unified": 3, "sampletype": 2}

BARCODE_COLS = ("corrected_barcode_fieldsample", "fieldsample_barcode",
                "barcode_fieldsample", "mfd_nr", "sample_id")
NATURA_COLS = ("naturnr", "habitat_typenumber", "habitat_typeno", "hab_code")
UNIFIED_RE = re.compile(r"(\d{3,5})")
EMPTY = {"", "none", "na", "n/a", "nan"}

VALID_SAMPLETYPE = {"soil", "water", "sediment"}
VALID_AREATYPE = {"natural", "urban", "subterranean", "agriculture"}

# --------------------------------------------------------------------------
# Value cleaning
# --------------------------------------------------------------------------


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def norm_barcode(value) -> str:
    """Normalize a barcode: uppercase, fix the MDF->MFD typo seen in some files."""
    s = clean(value).upper().replace(" ", "")
    if s.startswith("MDF"):
        s = "MFD" + s[3:]
    return s if BARCODE_RE.match(s) else ""


def norm_code(value) -> str:
    """Normalize a habitat code cell to a bare string ('2130', not '2130.0')."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def norm_term(value: str, valid: set[str]) -> str:
    """Title-case a sampletype/areatype, rejecting anything off-vocabulary.

    Handles legitimate combinations ('Soil/Sediment') and underscore-suffixed
    raw values ('natural_soil' -> ''  here, since that is a habitat field).
    """
    s = clean(value)
    if not s:
        return ""
    parts = [p for p in re.split(r"[/]", s) if p]
    cleaned = []
    for p in parts:
        token = p.strip().lower().split("_")[0]
        if token not in valid:
            return ""
        cleaned.append(token.title())
    return "/".join(cleaned)


# --------------------------------------------------------------------------
# Download / cache
# --------------------------------------------------------------------------


def fetch(rel_path: str, dest: Path) -> Path:
    """Download `rel_path` (under the repo's data/ dir) into the cache once."""
    if dest.exists():
        return dest
    dest.parent.mkdir(parents=True, exist_ok=True)
    url = RAW + "/" + urllib.parse.quote(rel_path)
    urllib.request.urlretrieve(url, dest)
    return dest


def download_all() -> dict:
    """Fetch the ontology, the NCBI join table, and all project files."""
    paths = {
        "ontology": fetch(ONTOLOGY_PATH, CACHE / "habitat_ontology.xlsx"),
        "ncbi_meta": fetch(NCBI_META_PATH, CACHE / "ncbi_meta.tsv"),
        "projects": {},
    }
    for proj, fnames in PROJECT_FILES.items():
        for fname in fnames:
            suffix = Path(fname).suffix
            dest = CACHE / proj / (Path(fname).stem.replace("/", "_") + suffix)
            try:
                fetch(f"metadata/{proj}/{fname}", dest)
                paths["projects"].setdefault(proj, []).append(dest)
            except Exception as exc:  # noqa: BLE001 - prototype: skip & report
                print(f"  WARN could not fetch {proj}/{fname}: {exc}")
    return paths


# --------------------------------------------------------------------------
# Habitat ontology -> code resolver
# --------------------------------------------------------------------------


def load_ontology(path: Path) -> dict:
    """Build a habitat-code -> (sampletype, areatype, hab1..3) index.

    Each ontology row is one habitat node; its terminal (deepest non-null) code
    plus its Natura-2000 code both resolve to that node's full lineage.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    code_index: dict[str, dict] = {}
    combos: set[tuple] = set()
    collisions = 0
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        node = {
            "mfd_sampletype": clean(row[0]), "mfd_areatype": clean(row[1]),
            "mfd_hab1": clean(row[3]), "mfd_hab2": clean(row[5]),
            "mfd_hab3": clean(row[7]),
        }
        combos.add(tuple(node.values()))
        terminal = norm_code(row[6]) or norm_code(row[4]) or norm_code(row[2])
        for code in {terminal, norm_code(row[8])}:
            if not code:
                continue
            if code in code_index and code_index[code] != node:
                collisions += 1
                continue
            code_index.setdefault(code, node)
    return {"code_index": code_index, "combos": combos, "collisions": collisions}


# --------------------------------------------------------------------------
# File reading (xlsx + delimited text), schema-driven record extraction
# --------------------------------------------------------------------------


def iter_tables(path: Path):
    """Yield (header: list[str], row_iter) for each sheet (xlsx) or the file (txt)."""
    if path.suffix.lower() in (".txt", ".csv", ".tsv"):
        text = path.read_text(encoding="latin-1")
        lines = text.splitlines()
        if not lines:
            return
        first = lines[0]
        delim = max(";", ",", "\t", key=first.count)
        reader = list(csv.reader(io.StringIO(text), delimiter=delim))
        yield [clean(c).lower() for c in reader[0]], reader[1:]
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            try:
                header = [clean(c).lower() for c in next(it)]
            except StopIteration:
                continue
            yield header, it


def parse_file(path: Path, project: str, ontology: dict) -> list[dict]:
    """Extract per-sample habitat records from one file (any schema)."""
    code_index = ontology["code_index"]
    records: list[dict] = []
    for header, row_iter in iter_tables(path):
        idx = {h: i for i, h in enumerate(header) if h}
        bc_col = next((idx[c] for c in BARCODE_COLS if c in idx), None)
        if bc_col is None:
            continue
        natura_col = next((idx[c] for c in NATURA_COLS if c in idx), None)
        has_standard = "mfd_hab1" in idx
        has_unified = "unified" in idx

        def cell(row, name):
            i = idx.get(name)
            return clean(row[i]) if i is not None and i < len(row) else ""

        empty_streak = 0
        for row in row_iter:
            if row is None or all(c in (None, "") for c in row):
                empty_streak += 1
                if empty_streak > 2000:  # bail past Excel dimension bloat
                    break
                continue
            empty_streak = 0
            barcode = norm_barcode(row[bc_col]) if bc_col < len(row) else ""
            if not barcode:
                continue

            st = norm_term(cell(row, "mfd_sampletype") or cell(row, "sample_type")
                           or cell(row, "type_sample") or cell(row, "sampletype"),
                           VALID_SAMPLETYPE)
            at = norm_term(cell(row, "mfd_areatype"), VALID_AREATYPE)

            rec = None
            # Tier 1: explicit mfd_hab1/2/3 with at least hab1 filled.
            if has_standard and cell(row, "mfd_hab1"):
                rec = {
                    "mfd_sampletype": st, "mfd_areatype": at,
                    "mfd_hab1": cell(row, "mfd_hab1"),
                    "mfd_hab2": cell(row, "mfd_hab2"),
                    "mfd_hab3": cell(row, "mfd_hab3"),
                    "source_signal": "standard", "source_code": "",
                }
            # Tier 2: resolve a Natura-2000 / MFD habitat code via the ontology.
            if rec is None and natura_col is not None and natura_col < len(row):
                node = code_index.get(norm_code(row[natura_col]))
                # consistency guard: reject a code whose resolved sampletype
                # contradicts the row's own sampletype (seen in P12_5).
                if node and (not st or node["mfd_sampletype"].lower() == st.lower()):
                    rec = {**node, "source_signal": "natura2000",
                           "source_code": norm_code(row[natura_col])}
            # Tier 3: agriculture `unified` code (so_a_NNNN).
            if rec is None and has_unified:
                m = UNIFIED_RE.search(cell(row, "unified"))
                node = code_index.get(m.group(1)) if m else None
                if node:
                    rec = {**node, "source_signal": "unified",
                           "source_code": m.group(1)}
            # Tier 4: only sample/area type recoverable.
            if rec is None:
                nature = cell(row, "nature_type") or cell(row, "habitattype")
                if not at and "natural" in nature.lower():
                    at = "Natural"
                if not at and "agri" in nature.lower():
                    at = "Agriculture"
                if not st and not at:
                    continue
                rec = {"mfd_sampletype": st, "mfd_areatype": at,
                       "mfd_hab1": "", "mfd_hab2": "", "mfd_hab3": "",
                       "source_signal": "sampletype", "source_code": ""}

            rec["mfd_sampletype"] = rec["mfd_sampletype"] or st
            rec["mfd_areatype"] = rec["mfd_areatype"] or at
            # habitat_type is a free-text submitter field ('marine', 'natural_soil')
            # -- a useful weak signal where mfd_hab1/2/3 were left blank.
            rec["habitat_type"] = (cell(row, "habitat_type")
                                   or cell(row, "habitattype"))
            rec["fieldsample_barcode"] = barcode
            rec["source_project"] = project
            rec["source_file"] = path.name
            rec["latitude"] = cell(row, "latitude") or cell(row, "lat")
            rec["longitude"] = cell(row, "longitude") or cell(row, "long")
            records.append(rec)
    return records


# --------------------------------------------------------------------------
# Normalize: collapse to one best record per barcode
# --------------------------------------------------------------------------


def hab_depth(rec: dict) -> int:
    return sum(populated(rec.get(f"mfd_hab{i}", "")) for i in (1, 2, 3))


def normalize(all_records: list[dict]) -> dict[str, dict]:
    """Keep the highest-signal record per barcode (ties: more hab depth wins)."""
    best: dict[str, dict] = {}
    for rec in all_records:
        bc = rec["fieldsample_barcode"]
        cur = best.get(bc)
        if cur is None:
            best[bc] = rec
            continue
        new_key = (SIGNAL_RANK.get(rec["source_signal"], 0), hab_depth(rec))
        cur_key = (SIGNAL_RANK.get(cur["source_signal"], 0), hab_depth(cur))
        if new_key > cur_key:
            best[bc] = rec
    return best


# --------------------------------------------------------------------------
# Output + coverage report
# --------------------------------------------------------------------------

COLUMNS = [
    "fieldsample_barcode", "mfd_sampletype", "mfd_areatype",
    "mfd_hab1", "mfd_hab2", "mfd_hab3", "habitat_type",
    "source_signal", "source_code", "source_project", "source_file",
    "latitude", "longitude",
]


def populated(value: str) -> bool:
    return bool(value) and str(value).lower() not in EMPTY


def write_table(table: dict[str, dict]) -> None:
    OUT_TSV.parent.mkdir(parents=True, exist_ok=True)
    lines = ["\t".join(COLUMNS)]
    for bc in sorted(table):
        rec = table[bc]
        lines.append("\t".join(str(rec.get(c, "")) for c in COLUMNS))
    OUT_TSV.write_text("\n".join(lines) + "\n")


def coverage_report(table: dict[str, dict], ncbi_meta: Path) -> dict:
    """Join the normalized table against the PRJNA1071982 NMDC biosamples."""
    inputs = json.loads(CURATION_INPUTS.read_text())
    bsm_to_mfdid: dict[str, str] = {}
    for bsm_id, payload in inputs["biosamples"].items():
        attrs = payload.get("attributes") or {}
        mfdid = norm_barcode(attrs.get("MFDID") or attrs.get("isolate") or "")
        if mfdid:
            bsm_to_mfdid[bsm_id] = mfdid

    nmdc_mfdids = set(bsm_to_mfdid.values())
    joinable = nmdc_mfdids & set(table)

    by_signal: dict[str, int] = {}
    hab1 = hab2 = hab3 = 0
    for mfdid in joinable:
        rec = table[mfdid]
        by_signal[rec["source_signal"]] = by_signal.get(rec["source_signal"], 0) + 1
        hab1 += populated(rec.get("mfd_hab1", ""))
        hab2 += populated(rec.get("mfd_hab2", ""))
        hab3 += populated(rec.get("mfd_hab3", ""))

    ncbi_mfdids = set()
    for ln in ncbi_meta.read_text().splitlines()[1:]:
        cols = ln.split("\t")
        if cols:
            bc = norm_barcode(cols[0])
            if bc:
                ncbi_mfdids.add(bc)

    # Re-evaluate the env-triad sentinels currently left unresolved. Buckets:
    #   resolvable_hab2 - rec has hab2 -> a precise ENVO term is derivable
    #   partial         - rec joined, no hab2, but sampletype/areatype/habitat_type
    #                     is present -> a coarser ENVO term is derivable
    #   joined_empty    - rec joined but carries no usable field
    #   not_joined      - barcode absent from the metadata repo entirely
    sentinel_recovered = {"env_broad_scale": 0, "env_local_scale": 0, "env_medium": 0}
    sentinel_total = {"env_broad_scale": 0, "env_local_scale": 0, "env_medium": 0}
    sentinel_status = {"resolvable_hab2": 0, "partial": 0,
                       "joined_empty": 0, "not_joined": 0}
    if CURATION_REPORT.exists():
        report = json.loads(CURATION_REPORT.read_text())
        for r in report["rows"]:
            if r["outcome"] != "left_sentinel":
                continue
            slot = r["slot"]
            sentinel_total[slot] = sentinel_total.get(slot, 0) + 1
            rec = table.get(bsm_to_mfdid.get(r["biosample_id"], ""))
            if rec is None:
                sentinel_status["not_joined"] += 1
            elif populated(rec.get("mfd_hab2", "")):
                sentinel_status["resolvable_hab2"] += 1
                sentinel_recovered[slot] = sentinel_recovered.get(slot, 0) + 1
            elif any(populated(rec.get(f, "")) for f in
                     ("mfd_hab1", "mfd_sampletype", "mfd_areatype", "habitat_type")):
                sentinel_status["partial"] += 1
            else:
                sentinel_status["joined_empty"] += 1

    return {
        "pinned_commit": REF,
        "habitat_table_barcodes": len(table),
        "nmdc_biosamples_with_mfdid": len(bsm_to_mfdid),
        "joinable_to_habitat_table": len(joinable),
        "joinable_pct": round(100 * len(joinable) / len(nmdc_mfdids), 1),
        "not_joinable": len(nmdc_mfdids - set(table)),
        "joinable_by_signal": dict(sorted(by_signal.items())),
        "joinable_with_hab1": hab1,
        "joinable_with_hab2": hab2,
        "joinable_with_hab3": hab3,
        "ncbi_meta_mfdids": len(ncbi_mfdids),
        "ncbi_meta_vs_inputs_match": len(ncbi_mfdids & nmdc_mfdids),
        "sentinels_total": sentinel_total,
        "sentinels_recoverable_via_hab2": sentinel_recovered,
        "sentinels_status": sentinel_status,
    }


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------


def main() -> None:
    print(f"Downloading mfd_metadata @ {REF[:10]} (cached in {CACHE}) ...")
    paths = download_all()

    ontology = load_ontology(paths["ontology"])
    print(f"Ontology: {len(ontology['code_index'])} resolvable habitat codes, "
          f"{len(ontology['combos'])} valid hab combos "
          f"({ontology['collisions']} code collisions skipped)")

    all_records: list[dict] = []
    n_files = 0
    for proj, files in paths["projects"].items():
        for path in files:
            n_files += 1
            try:
                all_records.extend(parse_file(path, proj, ontology))
            except Exception as exc:  # noqa: BLE001 - prototype: skip & report
                print(f"  WARN failed to parse {proj}/{path.name}: {exc}")
    print(f"Parsed {len(all_records)} habitat records from {n_files} files")

    table = normalize(all_records)
    write_table(table)
    print(f"Wrote {len(table)} unique barcodes -> {OUT_TSV.relative_to(REPO)}")

    report = coverage_report(table, paths["ncbi_meta"])
    OUT_REPORT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPORT.write_text(json.dumps(report, indent=2) + "\n")

    print("\n=== PRJNA1071982 coverage ===")
    print(f"  NMDC biosamples with an MFDID  : {report['nmdc_biosamples_with_mfdid']}")
    print(f"  joinable to habitat table      : {report['joinable_to_habitat_table']} "
          f"({report['joinable_pct']}%)")
    print(f"  not joinable                   : {report['not_joinable']}")
    print(f"  by signal tier                 : {report['joinable_by_signal']}")
    print(f"  joinable with hab1 / hab2 / hab3: {report['joinable_with_hab1']} / "
          f"{report['joinable_with_hab2']} / {report['joinable_with_hab3']}")
    print(f"  NCBI-meta cross-check          : "
          f"{report['ncbi_meta_vs_inputs_match']} / {report['ncbi_meta_mfdids']} match")
    print("\n=== env-triad sentinels ===")
    s = report["sentinels_status"]
    total = sum(s.values())
    print(f"  current sentinels (all slots)  : {total}")
    print(f"    resolvable via hab2          : {s['resolvable_hab2']}")
    print(f"    partial (coarser term only)  : {s['partial']}")
    print(f"    joined but empty             : {s['joined_empty']}")
    print(f"    not joinable (absent in repo): {s['not_joined']}")
    print(f"\nReport -> {OUT_REPORT.relative_to(REPO)}")


if __name__ == "__main__":
    sys.exit(main())
